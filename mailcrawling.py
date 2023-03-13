import os
import openpyxl as op
import win32com.client
import arrow

# -------------- Outlook Variations
# Beginning of Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
# The 'exture3' inbox in the total inbox(index: 6) 
inbox = outlook.GetDefaultFolder(6).Folders["exture3"]
# all messages in the 'exture3'
messages = inbox.Items
# all messages' count
msg_count = messages.count
# var for skip messages numbering
skip = 1

# -------------- Excel Variations
# Beginning of Excel
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = True
# Set the new workbook
wb = op.Workbook()
# Active the sheet
ws = wb.active

sites = []

# -------------- Class
class Site:
    def __init__(self, ws, idx):
        self.ws = ws
        self.idx = idx
        self.list = []
        # Add in the site list
        sites.append(self)
    
    # -------------- Mail Crawling Function
    def MailCrawling(self):
        # Viewing Progress
        ViewingProgress()
        
        # The arrow object to get the received time.
        arrowobj = arrow.get(mail.ReceivedTime)
        
        # Col A: Numbering
        self.ws[f'A{self.idx}'] = self.idx
        # Col B: Received datetime(set 'tzinfo' to 'None')
        self.ws[f'B{self.idx}'] = arrowobj.datetime.replace(tzinfo=None)
        # Col C: Sender Name(Sender Email Address)
        if mail.Class == 43 and mail.SenderEmailType == 'EX' and mail.Sender.GetExchangeUser() != None:
            self.ws[f'C{self.idx}'] = mail.SenderName + "(" + mail.Sender.GetExchangeUser().PrimarySmtpAddress + ")"
        else:
            self.ws[f'C{self.idx}'] = mail.SenderName + "(" + mail.SenderEmailAddress + ")"
        # Col D: receiving email addresse
        self.ws[f'D{self.idx}'] = mail.To
        # Col E: Mail Title
        self.ws[f'E{self.idx}'] = mail.Subject
        # Col F: Mail Body
        self.ws[f'F{self.idx}'] = mail.Body
        
        # Get the mail attachments
        attachments = mail.Attachments
        # Mail attachments' count
        r = attachments.count
        # Loop to save attachments (File Name: MailNumbering_AttachmentCount_AttachmentTitle)
        for j in range(1, r+1):
            attachment = attachments.Item(j)
            attachment.SaveASFile(AttachPath + str(self.idx) + "_" + str(j) + "_" + str(attachment)) # File Name
        return self.idx

# -------------- Site
kb = Site(wb.create_sheet('KB', 0), 1)
kb.list = ['KB', 'kb']

bnk = Site(wb.create_sheet('BNK', 1), 1)
bnk.list = ['BNK', 'bnk']

hana = Site(wb.create_sheet('하나', 2), 1)
hana.list = ['하나금융', '하나투자']

yt = Site(wb.create_sheet('유안타', 3), 1)
yt.list = ['유안타', 'yuanta']

eb = Site(wb.create_sheet('이베스트', 4), 1)
eb.list = ['이베스트', 'ebst']

hk = Site(wb.create_sheet('한투', 5), 1)
hk.list = ['한국투자', '한국']

sf = Site(wb.create_sheet('삼성선물', 6), 1)
sf.list = ['삼성선물', '삼성']

nhf = Site(wb.create_sheet('NH선물', 7), 1)
nhf.list = ['nh', 'NH', 'nh선물', 'NH선물']

hh = Site(wb.create_sheet('한화', 8), 1)
hh.list = ['한화', 'hanhwa']

# -------------- File Path
SavePath = "C:\\Users\\gee\\Desktop\\"
AttachPath = "C:\\Users\\gee\\Desktop\\attachments\\"

# -------------- Viewing Progress Function
def ViewingProgress():
    progress = 0
    
    os.system('cls')
    
    for site in sites:
        progress += site.idx-1
    
    print("In progress... (" + str(progress+skip) + "/" + str(msg_count)+ ")")

# -------------- MAIN
if __name__ == '__main__':
    # -------------- Mail Crawling Filtering Kewords
    for mail in messages:
        chk = 0
        for site in sites:
            if any(str in mail.Subject + mail.Body for str in site.list):
                site.idx = site.MailCrawling()
                site.idx += 1
                chk = 1
                break
        if chk == 0:
            ViewingProgress()
            skip += 1
            
    # -------------- Save Excel File
    wb.save(SavePath + "mailcrawling.xlsx")
    print("complete!")